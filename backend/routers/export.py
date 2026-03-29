# routers/export.py
# Export endpoints — class grades to Excel, paper results to PDF

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database.database import get_db
from models.models import (
    Class, Exam, Question, Student, StudentPaper,
    StudentAnswer, ClassEnrollment, PaperStatus
)
import io
from datetime import datetime

router = APIRouter(prefix="/api/export", tags=["Export"])


# ─── Export class grades to Excel ────────────────────────────────────────────

@router.get("/class/{class_id}/excel", summary="Export class grades to Excel")
def export_class_excel(class_id: int, db: Session = Depends(get_db)):
    """
    Exports all student grades for a class to an Excel file.
    One sheet per exam, plus a summary sheet showing all exams side by side.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    class_ = db.query(Class).filter(Class.id == class_id).first()
    if not class_:
        raise HTTPException(status_code=404, detail="Class not found")

    exams = db.query(Exam).filter(Exam.class_id == class_id).all()
    enrollments = db.query(ClassEnrollment).filter(
        ClassEnrollment.class_id == class_id
    ).all()
    students = [e.student for e in enrollments]

    wb = Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill   = PatternFill("solid", fgColor="2D6A4F")
    subhead_fill  = PatternFill("solid", fgColor="52B788")
    pass_fill     = PatternFill("solid", fgColor="D8F3DC")
    fail_fill     = PatternFill("solid", fgColor="FFE8E0")
    alt_fill      = PatternFill("solid", fgColor="F8F4EE")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    subhead_font  = Font(bold=True, color="FFFFFF", size=10)
    bold_font     = Font(bold=True, size=10)
    normal_font   = Font(size=10)
    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left",   vertical="center")
    thin_border   = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_header_row(ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = center_align
            cell.border = thin_border

    def style_cell(cell, fill=None, font=None, align=None):
        if fill:  cell.fill      = fill
        if font:  cell.font      = font
        if align: cell.alignment = align
        cell.border = thin_border

    # ── Summary Sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:Z1")
    title_cell = ws["A1"]
    title_cell.value     = f"{class_.name} — {class_.subject} — Grade Summary"
    title_cell.font      = Font(bold=True, size=14, color="2D6A4F")
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 28

    # Sub-title
    ws.merge_cells("A2:Z2")
    sub_cell = ws["A2"]
    sub_cell.value     = f"Exported: {datetime.now().strftime('%B %d, %Y')}"
    sub_cell.font      = Font(size=10, color="5A5A52", italic=True)
    sub_cell.alignment = center_align
    ws.row_dimensions[2].height = 18

    ws.append([])  # blank row

    # Build header: Student No | Name | Exam1 | Exam1% | Exam2 | ... | Overall Avg
    headers = ["Student No", "Student Name"]
    for exam in exams:
        headers += [exam.name, "%"]
    headers.append("Overall Average")
    ws.append(headers)
    style_header_row(ws, 4, len(headers))
    ws.row_dimensions[4].height = 20

    # Data rows
    for r_idx, student in enumerate(students, start=5):
        row_data   = [student.student_no, f"{student.first_name} {student.last_name}"]
        percentages = []

        for exam in exams:
            paper = db.query(StudentPaper).filter(
                StudentPaper.exam_id    == exam.id,
                StudentPaper.student_id == student.id,
                StudentPaper.status     == PaperStatus.graded,
            ).first()

            if paper and paper.total_score is not None and paper.max_score:
                pct = round(paper.total_score / paper.max_score * 100, 1)
                row_data += [f"{paper.total_score}/{paper.max_score}", f"{pct}%"]
                percentages.append(pct)
            else:
                row_data += ["—", "—"]

        avg = f"{round(sum(percentages)/len(percentages),1)}%" if percentages else "—"
        row_data.append(avg)
        ws.append(row_data)

        # Style row
        is_alt = (r_idx % 2 == 0)
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=r_idx, column=col)
            cell.font      = normal_font
            cell.alignment = center_align if col != 2 else left_align
            cell.border    = thin_border
            if col == 2:
                cell.font = bold_font
            if is_alt:
                cell.fill = alt_fill

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    for i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # ── Per-exam sheets ───────────────────────────────────────────────────────
    for exam in exams:
        safe_name = exam.name[:28].replace("/", "-").replace("\\", "-")
        ws2 = wb.create_sheet(title=safe_name)

        questions = db.query(Question).filter(
            Question.exam_id == exam.id
        ).order_by(Question.question_no).all()

        max_total = sum(q.max_score for q in questions)

        # Title
        ws2.merge_cells("A1:Z1")
        t = ws2["A1"]
        t.value     = f"{exam.name} — {class_.name}"
        t.font      = Font(bold=True, size=13, color="2D6A4F")
        t.alignment = center_align
        ws2.row_dimensions[1].height = 26
        ws2.append([])

        # Header
        headers2 = ["Student No", "Student Name"]
        for q in questions:
            headers2.append(f"Q{q.question_no}\n({q.max_score}pt)")
        headers2 += ["Total", "%", "Status"]
        ws2.append(headers2)
        style_header_row(ws2, 3, len(headers2))
        ws2.row_dimensions[3].height = 32

        for r_idx, student in enumerate(students, start=4):
            paper = db.query(StudentPaper).filter(
                StudentPaper.exam_id    == exam.id,
                StudentPaper.student_id == student.id,
            ).first()

            row2 = [student.student_no, f"{student.first_name} {student.last_name}"]

            if paper and paper.status == PaperStatus.graded:
                answers = {
                    a.question_id: a
                    for a in db.query(StudentAnswer).filter(
                        StudentAnswer.paper_id == paper.id
                    ).all()
                }
                for q in questions:
                    ans   = answers.get(q.id)
                    score = None
                    if ans:
                        score = ans.teacher_score if ans.teacher_score is not None else ans.score
                    row2.append(score if score is not None else "—")

                pct    = round(paper.total_score / max_total * 100, 1) if max_total else 0
                status = "PASSED" if pct >= 75 else "FAILED"
                row2  += [f"{paper.total_score}/{max_total}", f"{pct}%", status]
            else:
                row2 += ["—"] * len(questions)
                row2 += ["—", "—", "Not graded"]

            ws2.append(row2)

            # Style
            is_alt  = (r_idx % 2 == 0)
            is_pass = len(row2) >= 3 and row2[-1] == "PASSED"
            is_fail = len(row2) >= 3 and row2[-1] == "FAILED"

            for col in range(1, len(row2) + 1):
                cell = ws2.cell(row=r_idx, column=col)
                cell.font      = normal_font
                cell.alignment = center_align if col != 2 else left_align
                cell.border    = thin_border
                if col == 2:
                    cell.font = bold_font
                elif col == len(row2) and is_pass:
                    cell.fill = pass_fill
                    cell.font = Font(bold=True, size=10, color="2D6A4F")
                elif col == len(row2) and is_fail:
                    cell.fill = fail_fill
                    cell.font = Font(bold=True, size=10, color="C1440E")
                elif is_alt:
                    cell.fill = alt_fill

        ws2.column_dimensions["A"].width = 14
        ws2.column_dimensions["B"].width = 24
        for i in range(3, len(headers2) + 1):
            ws2.column_dimensions[get_column_letter(i)].width = 12

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{class_.name.replace(' ', '_')}_grades_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── Export single paper results to PDF ──────────────────────────────────────

@router.get("/paper/{paper_id}/pdf", summary="Export one student's results to PDF")
def export_paper_pdf(paper_id: int, db: Session = Depends(get_db)):
    """Generates a clean PDF report for one student's exam results."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    paper = db.query(StudentPaper).filter(StudentPaper.id == paper_id).first()
    if not paper:
        raise HTTPException(status_code=404, detail="Paper not found")

    exam      = db.query(Exam).filter(Exam.id == paper.exam_id).first()
    questions = {
        q.id: q for q in db.query(Question).filter(
            Question.exam_id == paper.exam_id
        ).all()
    }
    answers = db.query(StudentAnswer).filter(
        StudentAnswer.paper_id == paper_id
    ).all()

    GREEN      = colors.HexColor("#2D6A4F")
    LIGHT_GREEN= colors.HexColor("#52B788")
    CREAM      = colors.HexColor("#F8F4EE")
    DANGER     = colors.HexColor("#C1440E")
    MUTED      = colors.HexColor("#5A5A52")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
          leftMargin=2*cm, rightMargin=2*cm,
          topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    s_title  = ParagraphStyle('t', fontSize=20, textColor=GREEN,
                fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=4)
    s_sub    = ParagraphStyle('s', fontSize=11, textColor=MUTED,
                fontName='Helvetica', alignment=TA_CENTER, spaceAfter=2)
    s_h2     = ParagraphStyle('h2', fontSize=12, textColor=GREEN,
                fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=4)
    s_body   = ParagraphStyle('b', fontSize=10, textColor=colors.HexColor("#1A1A18"),
                fontName='Helvetica', spaceAfter=4, leading=14)
    s_label  = ParagraphStyle('l', fontSize=8, textColor=MUTED,
                fontName='Helvetica-Bold', spaceAfter=2)

    story = []

    # Header
    story.append(Paragraph("GrAid", s_title))
    story.append(Paragraph("Student Examination Report", s_sub))
    story.append(HRFlowable(width="100%", thickness=1.5, color=LIGHT_GREEN,
                            spaceBefore=6, spaceAfter=10))

    # Info table
    pct    = round(paper.total_score / paper.max_score * 100, 1) \
             if paper.max_score and paper.total_score is not None else 0
    status = "PASSED" if pct >= 75 else "FAILED"
    s_color = colors.HexColor("#2D6A4F") if pct >= 75 else DANGER

    info = [
        ["Student",  paper.student_name or "—",
         "Score",    f"{paper.total_score}/{paper.max_score}"],
        ["Exam",     exam.name if exam else "—",
         "Percentage", f"{pct}%"],
        ["Subject",  exam.subject if exam else "—",
         "Status",   status],
        ["Date",     paper.graded_at.strftime("%B %d, %Y") if paper.graded_at else "—",
         "", ""],
    ]
    t = Table(info, colWidths=[2.5*cm, 7*cm, 3*cm, 3*cm])
    t.setStyle(TableStyle([
        ('FONTNAME',  (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE',  (0,0), (-1,-1), 9.5),
        ('FONTNAME',  (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',  (2,0), (2,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0,0), (0,-1), MUTED),
        ('TEXTCOLOR', (2,0), (2,-1), MUTED),
        ('BACKGROUND',(0,0), (-1,-1), CREAM),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, CREAM]),
        ('GRID',      (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
        ('PADDING',   (0,0), (-1,-1), 6),
        ('TEXTCOLOR', (3,2), (3,2), s_color),
        ('FONTNAME',  (3,2), (3,2), 'Helvetica-Bold'),
    ]))
    story += [t, Spacer(1, 14)]

    # Answer breakdown
    story.append(Paragraph("Answer Breakdown", s_h2))
    story.append(HRFlowable(width="100%", thickness=0.5,
                            color=LIGHT_GREEN, spaceAfter=6))

    for ans in sorted(answers, key=lambda a: questions[a.question_id].question_no
                      if a.question_id in questions else 0):
        q = questions.get(ans.question_id)
        if not q: continue

        final = ans.teacher_score if ans.teacher_score is not None else ans.score
        pct_q = round(final / q.max_score * 100) if q.max_score and final is not None else 0
        is_correct = final is not None and final >= q.max_score
        row_color  = colors.HexColor("#D8F3DC") if is_correct else colors.HexColor("#FFE8E0")

        qdata = [
            [f"Q{q.question_no}  {q.question_type.replace('_',' ').title()}",
             f"{final if final is not None else '—'} / {q.max_score} pts",
             f"{pct_q}%"],
        ]
        qt = Table(qdata, colWidths=[10*cm, 4*cm, 2.5*cm])
        qt.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), row_color),
            ('FONTNAME',   (0,0), (0,0),  'Helvetica-Bold'),
            ('FONTNAME',   (1,0), (2,0),  'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 9.5),
            ('TEXTCOLOR',  (1,0), (2,0),
             colors.HexColor("#2D6A4F") if is_correct else DANGER),
            ('ALIGN',      (1,0), (2,0), 'RIGHT'),
            ('PADDING',    (0,0), (-1,-1), 6),
            ('ROUNDEDCORNERS', [4]),
        ]))
        story.append(qt)

        if ans.extracted_text:
            story.append(Paragraph(f"<b>OCR Read:</b> {ans.extracted_text}", s_body))
        story.append(Paragraph(f"<b>Answer Key:</b> {q.answer_key[:120]}", s_body))
        if ans.feedback:
            story.append(Paragraph(f"<b>Feedback:</b> {ans.feedback[:300]}", s_body))
        if ans.teacher_score is not None:
            story.append(Paragraph(
                f"<b>Teacher Override:</b> {ans.teacher_score}/{q.max_score}"
                + (f" — {ans.teacher_note}" if ans.teacher_note else ""),
                ParagraphStyle('ov', fontSize=9, textColor=colors.HexColor("#9a6700"),
                               fontName='Helvetica-Oblique', spaceAfter=4)
            ))
        story.append(Spacer(1, 6))

    # Footer
    story.append(HRFlowable(width="100%", thickness=0.5,
                            color=LIGHT_GREEN, spaceBefore=10, spaceAfter=6))
    story.append(Paragraph(
        f"Generated by GrAid · {datetime.now().strftime('%B %d, %Y %I:%M %p')}",
        ParagraphStyle('foot', fontSize=8, textColor=MUTED,
                       fontName='Helvetica', alignment=TA_CENTER)
    ))

    doc.build(story)
    buf.seek(0)

    student_name = (paper.student_name or "student").replace(" ", "_")
    filename     = f"{student_name}_{exam.name.replace(' ','_')}_results.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
